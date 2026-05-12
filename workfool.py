from openpyxl import load_workbook
import os

source_file = r"C:\Users\Zaki_Mir\Downloads\Book2.xlsx"
output_folder = r"C:\Users\Zaki_Mir\Documents\output"

os.makedirs(output_folder, exist_ok=True)

wb = load_workbook(source_file)

for sheet_name in wb.sheetnames:
    new_wb = load_workbook(source_file)
    
    a1_value = wb[sheet_name]['A1'].value
    file_name = str(a1_value).strip() if a1_value else sheet_name
    
    for char in r'\/:*?"<>|':
        file_name = file_name.replace(char, "")
    
    sheets_to_remove = [s for s in new_wb.sheetnames if s != sheet_name]
    for s in sheets_to_remove:
        del new_wb[s]
    
    output_path = os.path.join(output_folder, f"{file_name}.xlsx")
    new_wb.save(output_path)
    print(f"Saved: {output_path}")